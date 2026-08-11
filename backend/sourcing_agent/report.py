"""Turns the agent's findings into artifacts a human buyer can act on:
a JSON dump, a one-page HTML brief, and an .xlsx action file."""
from __future__ import annotations
import json
from dataclasses import asdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .schema import NormalizedQuote, BundleOpportunity, RiskFlag

INR = "\u20b9"


def _fmt(n):
    return f"{INR}{n:,.0f}"


def write_json(path: Path, quotes, bundles, risks):
    payload = {
        "normalized_quotes": [q.to_dict() for q in quotes],
        "bundle_opportunities": [asdict(b) for b in bundles],
        "risk_flags": [asdict(r) for r in risks],
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(path: Path, quotes, bundles, risks):
    wb = Workbook()
    hdr = Font(name="Arial", bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3A5F")
    base = Font(name="Arial")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left")

    ws = wb.active
    ws.title = "Bundling Opportunities"
    cols = ["Item", "Brands", "Combined Qty/mo", "Current Rs",
            "Best Vendor", "Best Rs", "Saving/unit", "Annual Saving Rs"]
    ws.append(cols)
    for b in bundles:
        ws.append([b.canonical_name, ", ".join(b.brands), b.combined_monthly_qty,
                   b.current_blended_price, b.best_vendor, b.best_price,
                   b.unit_saving, b.annual_saving])
    style_header(ws, len(cols))

    ws2 = wb.create_sheet("Risk Register")
    cols2 = ["Severity", "Item", "Risk Type", "Detail", "Recommended Action"]
    ws2.append(cols2)
    for r in risks:
        ws2.append([r.severity, r.canonical_name, r.risk_type,
                    r.detail, r.recommended_action])
    style_header(ws2, len(cols2))

    ws3 = wb.create_sheet("Normalized Quotes")
    cols3 = ["Canonical ID", "Item", "Vendor", "Unit Price", "UOM",
             "MOQ", "Lead (days)", "Valid Until", "Match Conf", "Raw Description"]
    ws3.append(cols3)
    for q in quotes:
        ws3.append([q.canonical_id, q.canonical_name, q.vendor, q.unit_price,
                    q.uom, q.moq, q.lead_time_days, q.quote_valid_until,
                    q.match_confidence, q.raw_description])
    style_header(ws3, len(cols3))

    for ws_ in (ws, ws2, ws3):
        for col in ws_.columns:
            width = max((len(str(c.value)) if c.value is not None else 0)
                        for c in col) + 2
            ws_.column_dimensions[col[0].column_letter].width = min(width, 48)
            for c in col:
                if c.row > 1 and not c.font.bold:
                    c.font = base
    wb.save(path)


def write_html(path: Path, quotes, bundles, risks):
    total = sum(b.annual_saving for b in bundles)
    sev_color = {"HIGH": "#c0392b", "MEDIUM": "#d68910", "LOW": "#7d8a99"}
    b_rows = "".join(
        f"<tr><td><b>{b.canonical_name}</b><br><span class=r>{b.rationale}</span></td>"
        f"<td>{', '.join(b.brands)}</td><td>{b.combined_monthly_qty:,.0f}</td>"
        f"<td>{INR}{b.current_blended_price:.2f}</td>"
        f"<td>{b.best_vendor}<br>{INR}{b.best_price:.2f}</td>"
        f"<td class=save>{_fmt(b.annual_saving)}</td></tr>" for b in bundles)
    r_rows = "".join(
        f"<tr><td><span class=sev style='background:{sev_color[r.severity]}'>"
        f"{r.severity}</span></td><td><b>{r.canonical_name}</b><br>{r.risk_type}</td>"
        f"<td>{r.detail}</td><td>{r.recommended_action}</td></tr>" for r in risks)

    html = f"""<!doctype html><meta charset=utf-8>
<title>Think9 Sourcing Agent — Brief</title>
<style>
 body{{font:15px/1.5 -apple-system,Segoe UI,Roboto,Arial;margin:0;background:#0f1620;color:#e8edf2}}
 .wrap{{max-width:960px;margin:0 auto;padding:32px}}
 h1{{font-size:26px;margin:0 0 4px}} .sub{{color:#8fa3b8;margin-bottom:24px}}
 .kpis{{display:flex;gap:16px;margin-bottom:28px;flex-wrap:wrap}}
 .kpi{{background:#18202c;border:1px solid #263242;border-radius:12px;padding:18px 22px;flex:1;min-width:170px}}
 .kpi .n{{font-size:28px;font-weight:700;color:#4ec9a4}} .kpi .l{{color:#8fa3b8;font-size:13px}}
 h2{{font-size:18px;margin:26px 0 10px;border-left:3px solid #4ec9a4;padding-left:10px}}
 table{{width:100%;border-collapse:collapse;background:#141b25;border-radius:10px;overflow:hidden}}
 th,td{{text-align:left;padding:10px 12px;border-bottom:1px solid #212b38;vertical-align:top;font-size:13.5px}}
 th{{background:#1f2a38;color:#c7d4e2;font-size:12px;text-transform:uppercase;letter-spacing:.03em}}
 .save{{color:#4ec9a4;font-weight:700;white-space:nowrap}}
 .r{{color:#8fa3b8;font-size:12px}}
 .sev{{color:#fff;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:700}}
 .foot{{color:#63758a;font-size:12px;margin-top:28px}}
</style>
<div class=wrap>
 <h1>Think9 Cross-Portfolio Sourcing Agent</h1>
 <div class=sub>Autonomous brief generated from {len(quotes)} normalized quotes across 5 vendor documents</div>
 <div class=kpis>
   <div class=kpi><div class=n>{_fmt(total)}</div><div class=l>Annual savings identified</div></div>
   <div class=kpi><div class=n>{len(bundles)}</div><div class=l>Bundling opportunities</div></div>
   <div class=kpi><div class=n>{len(risks)}</div><div class=l>Supply-risk flags</div></div>
   <div class=kpi><div class=n>{sum(1 for r in risks if r.severity=='HIGH')}</div><div class=l>High-severity risks</div></div>
 </div>
 <h2>Volume-bundling opportunities</h2>
 <table><tr><th>Item &amp; rationale</th><th>Brands</th><th>Qty/mo</th><th>Current</th><th>Best vendor</th><th>Annual saving</th></tr>{b_rows}</table>
 <h2>Supply-risk register</h2>
 <table><tr><th>Severity</th><th>Item / type</th><th>Detail</th><th>Recommended action</th></tr>{r_rows}</table>
 <div class=foot>Every figure computed from parsed vendor data. Requires human buyer approval before any PO — see human-in-the-loop console.</div>
</div>"""
    path.write_text(html, encoding="utf-8")
